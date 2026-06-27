"""
Expenses & Petty Cash — HR Pro Module
Smart Paste system for petty cash tracking, employee advances,
expenses, received funds, and transfer fees.
"""
import re
import streamlit as st
import pandas as pd
from datetime import date, datetime
from database.db import get_db
from database.models import Employee, Advance, Expense, ReceivedFund
from components.cards import (
    section_card, empty_state, success_toast,
    danger_toast, warning_toast, kpi_card,
)
from components.layout import render_topbar
from components.theme import COLORS


# ─────────────────────────────────────────────────────────────────
#  SESSION-STATE HELPERS
# ─────────────────────────────────────────────────────────────────
#
# NOTE: expenses and received-funds used to live entirely in
# st.session_state (exp_expenses / exp_received_funds), which meant
# every refresh, browser close, or Streamlit rerun-after-error wiped
# them out. They are now persisted in the Expense / ReceivedFund tables
# (see database/models.py) and loaded fresh from the DB wherever they're
# needed — see _load_ledger() / _load_received_funds() below.
# exp_paste_result is the one piece that legitimately stays in
# session_state: it's just a transient "here's what the last Smart
# Paste run did" summary for the UI, not data — the underlying records
# it describes are already safely in the database by the time it's
# shown.

def _ensure_state():
    defaults = {
        "exp_paste_result":   None,
        "exp_filter":         "الكل",
        "exp_filter_emp":     None,
        "exp_filter_cat":     None,
        "exp_date_from":      date.today().replace(day=1),
        "exp_date_to":        date.today(),
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


CATEGORIES = [
    "وقود", "مواصلات", "أكل وشرب", "مواد",
    "معدات", "صيانة", "مكتبية", "أخرى",
]

CAT_MAP = {
    "بنزين": "وقود", "وقود": "وقود", "ديزل": "وقود",
    "فطار": "أكل وشرب", "غداء": "أكل وشرب", "عشاء": "أكل وشرب",
    "اكل": "أكل وشرب", "أكل": "أكل وشرب", "طعام": "أكل وشرب",
    "مياه": "أكل وشرب", "مشروبات": "أكل وشرب",
    "معدات": "معدات", "ادوات": "معدات",
    "مواد": "مواد", "خامات": "مواد",
    "صيانة": "صيانة", "اصلاح": "صيانة",
    "مواصلات": "مواصلات", "تاكسي": "مواصلات", "نقل": "مواصلات",
    "مكتبية": "مكتبية", "قرطاسية": "مكتبية",
}

TRANSFER_TAX_RATE = 0.002


# ─────────────────────────────────────────────────────────────────
#  ARABIC / NUMBER NORMALIZATION
# ─────────────────────────────────────────────────────────────────

_AR_DIGITS = str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789")
_DATE_RE   = re.compile(
    r'\b(\d{1,2})[/\-\\.](\d{1,2})(?:[/\-\\.](\d{2,4}))?\b'
)
_NUM_RE    = re.compile(r'(\d+(?:[.,]\d+)?)\s*(?:ج\.?م\.?|جنيه|LE|EGP)?')

ADVANCE_KEYWORDS = re.compile(r'سلف[هة]', re.UNICODE)

# Separators between items in a line (Arabic colon or spaces)
_SEP_PATTERN = re.compile(r'[,:؛;|/\\]+')


def _normalize_num(s: str) -> float:
    try:
        return float(s.replace(",", ".").strip())
    except ValueError:
        return 0.0


def _normalize_name(name: str) -> str:
    """Strip prefixes for fuzzy matching."""
    name = name.strip()
    for p in ["م. ", "م ", "أ. ", "أ ", "د. ", "د "]:
        if name.startswith(p):
            name = name[len(p):]
    return name.strip()


def _find_employee(raw_name: str, employees):
    """Smart fuzzy match against active employees."""
    norm = _normalize_name(raw_name)
    # Exact
    for e in employees:
        if _normalize_name(e.name) == norm:
            return e
    # Substring
    for e in employees:
        en = _normalize_name(e.name)
        if norm in en or en in norm:
            return e
    # Token overlap
    tokens = set(norm.split())
    for e in employees:
        et = set(_normalize_name(e.name).split())
        if tokens & et:
            return e
    return None


# ─────────────────────────────────────────────────────────────────
#  SMART PASTE PARSER
# ─────────────────────────────────────────────────────────────────

def _guess_category(desc: str) -> str:
    dl = desc.strip().lower()
    for kw, cat in CAT_MAP.items():
        if kw in dl:
            return cat
    return "أخرى"


def _parse_item_token(token: str):
    """
    Parse a single expense item token.
    Formats:
      سلفة فارس : 500
      سلفه فارس 500
      بنزين 250
      فطار عمال : 180
      فارس سلفة 500
    Returns dict with keys: is_advance, desc, emp_name_hint, amount
    """
    token = token.translate(_AR_DIGITS).strip()
    if not token:
        return None

    # Extract the amount (last or only number)
    nums = _NUM_RE.findall(token)
    if not nums:
        return None
    amount_str = nums[-1]
    amount = _normalize_num(amount_str)
    if amount <= 0:
        return None

    # Remove the amount from the text to get description
    text_part = _NUM_RE.sub("", token).strip()
    # Clean leftover punctuation
    text_part = re.sub(r'[:\-_،,;؛|/\\]+', ' ', text_part).strip()
    text_part = re.sub(r'\s+', ' ', text_part).strip()

    is_advance = bool(ADVANCE_KEYWORDS.search(text_part))

    # Strip the advance keyword to get the name
    emp_hint = ADVANCE_KEYWORDS.sub("", text_part).strip() if is_advance else None

    return {
        "is_advance": is_advance,
        "desc":       text_part,
        "emp_hint":   emp_hint if emp_hint else None,
        "amount":     amount,
    }


def _parse_paste_block(text: str, block_date: date, employees):
    """
    Parse one date block's lines.
    Lines can contain multiple items separated by spaces/keywords.
    """
    advances  = []
    expenses  = []
    unmatched = []

    # Split into potential items by looking for repeated structures
    # Strategy: scan for advance keywords and then split the rest
    # We tokenize by splitting on known delimiters between items
    raw = text.strip()
    raw = raw.translate(_AR_DIGITS)

    # Use a greedy approach: split on patterns like  "  word1 word2 : 123 "
    # We'll use a regex to find all (description : amount) pairs
    # Pattern: text optionally ending with ":" then a number
    item_pattern = re.compile(
        r'(?:سلف[هة]\s*[\w\s]+?\s*[:：]?\s*\d+(?:[.,]\d+)?|'  # advance pattern
        r'[\u0600-\u06FF\w]+(?:\s+[\u0600-\u06FF\w]+)*\s*[:：\s]\s*\d+(?:[.,]\d+)?)',
        re.UNICODE
    )
    # Simpler: split line into chunks around numbers
    # We'll tokenize differently — find all (text, number) pairs
    pair_re = re.compile(
        r'((?:سلف[هة]\s+)?[\u0600-\u06FFa-zA-Z][\u0600-\u06FFa-zA-Z\s]*?)'
        r'\s*[:：]?\s*'
        r'(\d+(?:[.,]\d+)?)\s*(?:ج\.?م\.?|جنيه|LE|EGP)?',
        re.UNICODE
    )

    items_found = pair_re.findall(raw)

    if not items_found:
        # Fallback: split by slash/pipe separators
        parts = _SEP_PATTERN.split(raw)
        for p in parts:
            parsed = _parse_item_token(p)
            if parsed:
                items_found.append((parsed["desc"], str(parsed["amount"])))

    for text_part, amount_str in items_found:
        text_part = text_part.strip()
        amount    = _normalize_num(amount_str)
        if amount <= 0 or not text_part:
            continue

        is_advance = bool(ADVANCE_KEYWORDS.search(text_part))
        emp_hint   = ADVANCE_KEYWORDS.sub("", text_part).strip() if is_advance else None

        # Clean up emp_hint
        if emp_hint:
            emp_hint = re.sub(r'[:\-،,;؛|/\\]+', ' ', emp_hint).strip()
            emp_hint = re.sub(r'\s+', ' ', emp_hint).strip()

        if is_advance and emp_hint:
            emp = _find_employee(emp_hint, employees)
            if emp:
                advances.append({
                    "date":        block_date,
                    "employee_id": emp.id,
                    "emp_name":    emp.name,
                    "amount":      amount,
                    "notes":       f"سلفة - Smart Paste {block_date}",
                    "matched":     True,
                })
            else:
                unmatched.append({
                    "date":     block_date,
                    "emp_hint": emp_hint,
                    "amount":   amount,
                    "reason":   "لم يُعثر على الموظف",
                })
        else:
            desc = text_part
            desc = re.sub(r'[:\-،,]+$', '', desc).strip()
            expenses.append({
                "date":     block_date,
                "desc":     desc if desc else "مصروف",
                "category": _guess_category(desc),
                "amount":   amount,
                "notes":    "",
            })

    return advances, expenses, unmatched


def parse_smart_paste(text: str, employees):
    """
    Full parser for the Smart Paste input.
    Handles multiple date blocks.
    """
    all_advances  = []
    all_expenses  = []
    all_unmatched = []

    date_re = re.compile(
        r'^(\d{1,2})[/\-\\.](\d{1,2})(?:[/\-\\.](\d{2,4}))?$'
    )

    lines = [l.strip() for l in text.strip().splitlines()]
    current_date  = date.today()
    current_lines = []

    def flush():
        if not current_lines:
            return
        block_text = " ".join(current_lines)
        a, e, u = _parse_paste_block(block_text, current_date, employees)
        all_advances.extend(a)
        all_expenses.extend(e)
        all_unmatched.extend(u)

    for line in lines:
        line_n = line.translate(_AR_DIGITS).strip()
        m = date_re.match(line_n)
        if m:
            flush()
            current_lines = []
            d, mo, yr = m.group(1), m.group(2), m.group(3)
            try:
                y = int(yr) if yr else date.today().year
                if y < 100:
                    y += 2000
                current_date = date(y, int(mo), int(d))
            except ValueError:
                current_date = date.today()
        elif line_n:
            current_lines.append(line_n)

    flush()

    return all_advances, all_expenses, all_unmatched


# ─────────────────────────────────────────────────────────────────
#  DATABASE: SAVE / LOAD
# ─────────────────────────────────────────────────────────────────

def _save_advance(employee_id: int, adv_date: date, amount: float, notes: str):
    db = get_db()
    try:
        db.add(Advance(
            employee_id=employee_id,
            date=adv_date,
            amount=amount,
            notes=notes,
        ))
        db.commit()
        return True, None
    except Exception as e:
        db.rollback()
        return False, str(e)
    finally:
        db.close()


def _save_expense(exp_date: date, desc: str, category: str, amount: float, notes: str = ""):
    db = get_db()
    try:
        db.add(Expense(
            date=exp_date,
            description=(desc or "مصروف").strip() or "مصروف",
            category=category or "أخرى",
            amount=amount,
            notes=notes or "",
        ))
        db.commit()
        return True, None
    except Exception as e:
        db.rollback()
        return False, str(e)
    finally:
        db.close()


def _save_received_fund(recv_date: date, amount: float, tax: float, net: float,
                         source: str, notes: str = ""):
    db = get_db()
    try:
        db.add(ReceivedFund(
            date=recv_date,
            amount=amount,
            tax=tax,
            net_amount=net,
            source=(source or "—").strip() or "—",
            notes=notes or "",
        ))
        db.commit()
        return True, None
    except Exception as e:
        db.rollback()
        return False, str(e)
    finally:
        db.close()


def _expense_rows_as_dicts(db) -> list:
    rows = db.query(Expense).order_by(Expense.id).all()
    return [{
        "kind":       "expense",
        "id":         r.id,
        "date":       r.date,
        "desc":       r.description,
        "category":   r.category or "أخرى",
        "amount":     r.amount,
        "notes":      r.notes or "",
        "is_advance": False,
    } for r in rows]


def _advance_rows_as_expense_dicts(db) -> list:
    """
    Advances live in their own table (Advance) since they're tied into
    payroll — but the Expenses page has always shown them merged into
    one unified "expenses" ledger alongside regular petty-cash expenses
    for reporting/balance purposes. This rebuilds that merged view from
    the DB instead of from a session-state list.
    """
    rows = (
        db.query(Advance, Employee.name)
        .outerjoin(Employee, Advance.employee_id == Employee.id)
        .order_by(Advance.id)
        .all()
    )
    out = []
    for adv, emp_name in rows:
        name = emp_name or "—"
        out.append({
            "kind":        "advance",
            "id":          adv.id,
            "employee_id": adv.employee_id,
            "date":        adv.date,
            "desc":        f"سلفة — {name}",
            "emp_name":    name,
            "category":    None,
            "amount":      adv.amount,
            "notes":       adv.notes or "",
            "is_advance":  True,
        })
    return out


def _load_ledger() -> list:
    """Unified, DB-backed list of expense dicts: real expenses + advances."""
    db = get_db()
    try:
        items = _expense_rows_as_dicts(db) + _advance_rows_as_expense_dicts(db)
    finally:
        db.close()
    items.sort(key=lambda e: (e["date"], e["id"]))
    return items


def _load_received_funds() -> list:
    db = get_db()
    try:
        rows = db.query(ReceivedFund).order_by(ReceivedFund.id).all()
        return [{
            "id":     r.id,
            "date":   r.date,
            "amount": r.amount,
            "tax":    r.tax,
            "net":    r.net_amount,
            "source": r.source or "—",
            "notes":  r.notes or "",
        } for r in rows]
    finally:
        db.close()


# ─────────────────────────────────────────────────────────────────
#  MAIN PAGE
# ─────────────────────────────────────────────────────────────────

def show():
    _ensure_state()
    render_topbar("المصروفات والخزينة الصغيرة", "💵", "المصروفات")

    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "📋 لوحة التحكم",
        "⚡ الإدخال الذكي",
        "💰 صندوق الوارد",
        "📂 السجلات والفلاتر",
        "📊 التقارير والتصدير",
    ])

    with tab1:
        _tab_dashboard()
    with tab2:
        _tab_smart_paste()
    with tab3:
        _tab_received_funds()
    with tab4:
        _tab_records()
    with tab5:
        _tab_reports()


# ─────────────────────────────────────────────────────────────────
#  TAB 1 — DASHBOARD
# ─────────────────────────────────────────────────────────────────

def _tab_dashboard():
    section_card("ملخص مالي", "📊", COLORS["primary"])

    expenses       = _load_ledger()
    received_funds = _load_received_funds()

    total_received = sum(r["amount"] for r in received_funds)
    total_tax      = sum(r["amount"] * TRANSFER_TAX_RATE for r in received_funds)
    net_received   = total_received - total_tax

    adv_total  = sum(e["amount"] for e in expenses if e.get("is_advance"))
    exp_total  = sum(e["amount"] for e in expenses if not e.get("is_advance"))
    total_exp  = adv_total + exp_total
    balance    = net_received - total_exp

    # Row 1
    c1, c2, c3, c4 = st.columns(4)
    kpi_card(c1, "💵", "إجمالي الوارد",
             f"{total_received:,.0f} ج.م", COLORS["success"])
    kpi_card(c2, "🏦", "رسوم التحويل",
             f"{total_tax:,.2f} ج.م", COLORS["warning"])
    kpi_card(c3, "✅", "صافي الوارد",
             f"{net_received:,.0f} ج.م", COLORS["info"])
    kpi_card(c4, "💰", "إجمالي السلف للموظفين",
             f"{adv_total:,.0f} ج.م", COLORS["accent"])

    st.markdown("<div style='height:14px;'></div>", unsafe_allow_html=True)

    # Row 2
    c5, c6, c7, _ = st.columns(4)
    kpi_card(c5, "🧾", "مصروفات أخرى",
             f"{exp_total:,.0f} ج.م", COLORS["secondary"])
    kpi_card(c6, "📉", "إجمالي المصروفات",
             f"{total_exp:,.0f} ج.م", COLORS["danger"])
    balance_color = COLORS["success"] if balance >= 0 else COLORS["danger"]
    kpi_card(c7, "🏧", "الرصيد المتبقي",
             f"{balance:,.0f} ج.م", balance_color,
             trend="✅ إيجابي" if balance >= 0 else "⚠️ سالب")

    # Formula note
    st.markdown(f"""
<div style="background:{COLORS['surface']};border:1px solid {COLORS['border']};
            border-radius:10px;padding:14px 20px;margin-top:18px;font-size:13px;
            color:{COLORS['text_muted']};">
  📐 <strong style="color:{COLORS['text']};">المعادلة:</strong>
  الرصيد = صافي الوارد − (إجمالي السلف + إجمالي المصروفات)
  = <strong style="color:{COLORS['primary']};">{net_received:,.0f}</strong> −
    <strong style="color:{COLORS['danger']};">{total_exp:,.0f}</strong>
  = <strong style="color:{balance_color};">{balance:,.0f} ج.م</strong>
</div>
""", unsafe_allow_html=True)

    # Recent activity
    if expenses or received_funds:
        st.markdown("<div style='height:20px;'></div>", unsafe_allow_html=True)
        section_card("آخر العمليات", "🕐", COLORS["primary_light"])
        all_txns = []
        for r in received_funds[-5:]:
            all_txns.append({
                "التاريخ": str(r["date"]),
                "النوع": "📥 وارد",
                "البيان": r.get("source", "—"),
                "المبلغ": f"+{r['amount']:,.0f}",
            })
        for e in expenses[-5:]:
            tag = "💰 سلفة" if e.get("is_advance") else "🧾 مصروف"
            all_txns.append({
                "التاريخ": str(e["date"]),
                "النوع": tag,
                "البيان": e.get("desc", e.get("emp_name", "—")),
                "المبلغ": f"-{e['amount']:,.0f}",
            })
        if all_txns:
            df = pd.DataFrame(all_txns[-10:])
            st.dataframe(df, use_container_width=True, hide_index=True)


# ─────────────────────────────────────────────────────────────────
#  TAB 2 — SMART PASTE
# ─────────────────────────────────────────────────────────────────

def _tab_smart_paste():
    section_card("الإدخال الذكي — Smart Paste", "⚡", COLORS["primary"])

    st.markdown(f"""
<div style="background:{COLORS['primary']}0d;border:1px solid {COLORS['primary']}22;
            border-radius:10px;padding:14px 18px;margin-bottom:18px;font-size:13px;
            color:{COLORS['text']};line-height:2;">
  <strong>📌 مثال على الصيغة:</strong><br>
  <code style="font-family:monospace;background:#f4f8f6;padding:2px 6px;border-radius:4px;direction:ltr;display:inline-block;">
  24/06/2026<br>
  سلفة فارس : 500  سلفة احمد : 300  بنزين : 250  فطار عمال : 180  مياه : 50<br>
  25/06/2026<br>
  سلفه فارس 200  معدات : 400
  </code><br><br>
  <strong>✅ الصيغ المدعومة:</strong>
  <code>سلفة فارس 500</code> &nbsp;|&nbsp;
  <code>سلفة فارس : 500</code> &nbsp;|&nbsp;
  <code>سلفه فارس 500ج</code> &nbsp;|&nbsp;
  <code>فارس سلفة 500</code>
</div>
""", unsafe_allow_html=True)

    paste_text = st.text_area(
        "📋 الصق النص هنا",
        height=160,
        placeholder="24/06/2026\nسلفة فارس : 500  بنزين : 250  فطار عمال : 180",
        key="exp_paste_input",
    )

    col_btn, col_clear = st.columns([3, 1])
    process = col_btn.button("⚡ معالجة وحفظ", type="primary", use_container_width=True)
    clear   = col_clear.button("🗑️ مسح", use_container_width=True)

    if clear:
        st.session_state.exp_paste_result = None
        st.rerun()

    if process and paste_text.strip():
        db = get_db()
        try:
            employees = db.query(Employee).filter(Employee.status == "Active").all()
        finally:
            db.close()

        advances, expenses, unmatched = parse_smart_paste(paste_text, employees)

        # Save advances to DB (Advance table)
        saved_adv = []
        failed_adv = []
        for adv in advances:
            ok, err = _save_advance(
                adv["employee_id"], adv["date"], adv["amount"], adv["notes"]
            )
            if ok:
                saved_adv.append(adv)
            else:
                failed_adv.append({**adv, "error": err})

        # Save expenses to DB (Expense table) — same per-item handling as
        # advances, so one bad row doesn't drop the rest of the paste.
        saved_exp = []
        failed_exp = []
        for e in expenses:
            ok, err = _save_expense(e["date"], e["desc"], e["category"], e["amount"], e.get("notes", ""))
            if ok:
                saved_exp.append(e)
            else:
                failed_exp.append({**e, "error": err})

        st.session_state.exp_paste_result = {
            "advances":   saved_adv,
            "expenses":   saved_exp,
            "unmatched":  unmatched,
            "failed_adv": failed_adv,
            "failed_exp": failed_exp,
        }
        st.rerun()

    # Show results
    result = st.session_state.get("exp_paste_result")
    if result:
        _show_paste_result(result)

    # Manual entry fallback
    with st.expander("✏️ إضافة مصروف يدوي", expanded=False):
        _manual_expense_form()


def _show_paste_result(result):
    advances   = result["advances"]
    expenses   = result["expenses"]
    unmatched  = result["unmatched"]
    failed     = result["failed_adv"]
    failed_exp = result.get("failed_exp", [])

    st.markdown("<div style='height:16px;'></div>", unsafe_allow_html=True)
    section_card("نتيجة المعالجة", "📊", COLORS["success"])

    c1, c2, c3, c4 = st.columns(4)
    kpi_card(c1, "💰", "سلف محفوظة",     str(len(advances)),  COLORS["success"])
    kpi_card(c2, "🧾", "مصروفات محفوظة", str(len(expenses)),  COLORS["primary"])
    kpi_card(c3, "⚠️", "غير معروف",      str(len(unmatched)), COLORS["warning"])
    kpi_card(c4, "❌", "عمليات فاشلة",   str(len(failed) + len(failed_exp)), COLORS["danger"])

    if advances:
        st.markdown("**✅ سلف تم حفظها في نظام السلف:**")
        df = pd.DataFrame([{
            "الموظف": a["emp_name"],
            "التاريخ": str(a["date"]),
            "المبلغ": f"{a['amount']:,.0f} ج.م",
        } for a in advances])
        st.dataframe(df, use_container_width=True, hide_index=True)

    if expenses:
        st.markdown("**🧾 مصروفات تم تسجيلها:**")
        df = pd.DataFrame([{
            "البيان":  e["desc"],
            "الفئة":   e["category"],
            "التاريخ": str(e["date"]),
            "المبلغ":  f"{e['amount']:,.0f} ج.م",
        } for e in expenses])
        st.dataframe(df, use_container_width=True, hide_index=True)

    if unmatched:
        st.warning("⚠️ الأسماء التالية لم تُطابق أي موظف:")
        for u in unmatched:
            st.markdown(
                f"- **{u['emp_hint']}** — {u['amount']:,.0f} ج.م  _(السبب: {u['reason']})_"
            )

    if failed:
        st.error("❌ بعض السلف لم تُحفظ:")
        for f in failed:
            st.markdown(f"- {f['emp_name']} — {f.get('error','خطأ غير معروف')}")

    if failed_exp:
        st.error("❌ بعض المصروفات لم تُحفظ:")
        for f in failed_exp:
            st.markdown(f"- {f.get('desc','—')} — {f.get('error','خطأ غير معروف')}")


def _manual_expense_form():
    with st.form("manual_exp_form", clear_on_submit=True):
        col1, col2 = st.columns(2)
        exp_date = col1.date_input("📅 التاريخ", value=date.today())
        amount   = col2.number_input("💰 المبلغ (ج.م)", min_value=0.0, step=10.0)
        desc     = st.text_input("📝 البيان", placeholder="بنزين، فطار عمال، مياه...")
        col3, col4 = st.columns(2)
        cat  = col3.selectbox("📂 الفئة", CATEGORIES)
        notes = col4.text_input("📎 ملاحظات", placeholder="اختياري")
        submitted = st.form_submit_button("💾 حفظ المصروف", type="primary", use_container_width=True)
        if submitted:
            if amount <= 0 or not desc.strip():
                st.error("يرجى إدخال البيان والمبلغ.")
            else:
                ok, err = _save_expense(exp_date, desc.strip(), cat, amount, notes)
                if ok:
                    success_toast("✅ تم حفظ المصروف.")
                    st.rerun()
                else:
                    danger_toast(f"خطأ في حفظ المصروف: {err}")


# ─────────────────────────────────────────────────────────────────
#  TAB 3 — RECEIVED FUNDS
# ─────────────────────────────────────────────────────────────────

def _tab_received_funds():
    section_card("إضافة وارد / صندوق", "📥", COLORS["success"])

    with st.form("recv_fund_form", clear_on_submit=True):
        col1, col2 = st.columns(2)
        recv_date  = col1.date_input("📅 التاريخ", value=date.today())
        amount     = col2.number_input("💵 المبلغ المستلم (ج.م)", min_value=0.0, step=100.0)
        source     = st.text_input("🏦 المصدر", placeholder="تحويل بنكي، نقدي،...")
        notes      = st.text_input("📎 ملاحظات", placeholder="اختياري")
        submitted  = st.form_submit_button("💾 تسجيل الوارد", type="primary", use_container_width=True)

        if submitted:
            if amount <= 0:
                st.error("المبلغ يجب أن يكون أكبر من صفر.")
            else:
                tax    = round(amount * TRANSFER_TAX_RATE, 2)
                net    = amount - tax
                ok, err = _save_received_fund(
                    recv_date, amount, tax, net, source.strip() or "—", notes.strip()
                )
                if ok:
                    success_toast(f"✅ تم تسجيل {amount:,.0f} ج.م — رسوم: {tax:,.2f} ج.م — صافي: {net:,.2f} ج.م")
                    st.rerun()
                else:
                    danger_toast(f"خطأ في حفظ الوارد: {err}")

    # Summary of received
    funds = _load_received_funds()
    if funds:
        st.markdown("<div style='height:16px;'></div>", unsafe_allow_html=True)
        section_card("سجل الواردات", "📋", COLORS["primary"])

        total_r = sum(f["amount"] for f in funds)
        total_t = sum(f["tax"]    for f in funds)
        total_n = sum(f["net"]    for f in funds)

        c1, c2, c3 = st.columns(3)
        kpi_card(c1, "💵", "إجمالي الوارد",  f"{total_r:,.0f} ج.م", COLORS["success"])
        kpi_card(c2, "🏦", "رسوم التحويل",   f"{total_t:,.2f} ج.م", COLORS["warning"])
        kpi_card(c3, "✅", "صافي الوارد",    f"{total_n:,.0f} ج.م", COLORS["info"])

        st.markdown("<div style='height:12px;'></div>", unsafe_allow_html=True)

        df = pd.DataFrame([{
            "التاريخ":            str(f["date"]),
            "المبلغ المستلم":     f"{f['amount']:,.0f}",
            "رسوم التحويل":       f"{f['tax']:,.2f}",
            "الصافي بعد الرسوم":  f"{f['net']:,.2f}",
            "المصدر":             f["source"],
            "ملاحظات":            f["notes"],
        } for f in funds])

        st.dataframe(df, use_container_width=True, hide_index=True)

        # Delete last
        if st.button("🗑️ حذف آخر وارد", key="del_last_recv"):
            db = get_db()
            try:
                last = db.query(ReceivedFund).order_by(ReceivedFund.id.desc()).first()
                if last:
                    db.delete(last)
                    db.commit()
            finally:
                db.close()
            st.rerun()
    else:
        empty_state("لا توجد واردات مسجلة", "💵")

    # Info box about transfer tax
    st.markdown(f"""
<div style="background:{COLORS['warning']}12;border:1px solid {COLORS['warning']}33;
            border-radius:10px;padding:14px 18px;margin-top:18px;font-size:13px;
            color:{COLORS['text']};">
  🏦 <strong>رسوم التحويل:</strong>
  لكل وارد تحويل بنكي، يُحسب: <strong>المبلغ × 0.002</strong><br>
  مثال: 10,000 ج.م × 0.002 = <strong>20 ج.م</strong> رسوم ←
  الصافي = <strong>9,980 ج.م</strong>
</div>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────
#  TAB 4 — RECORDS & FILTERS
# ─────────────────────────────────────────────────────────────────

def _tab_records():
    section_card("السجلات والفلاتر", "📂", COLORS["primary"])

    expenses = _load_ledger()
    if not expenses:
        empty_state("لا توجد مصروفات مسجلة بعد", "🧾")
        return

    # Filters
    col1, col2, col3, col4 = st.columns([2, 2, 2, 2])

    filter_type = col1.selectbox(
        "🔍 نوع العملية",
        ["الكل", "سلف فقط", "مصروفات فقط"],
    )
    all_cats = sorted({e.get("category", "أخرى") for e in expenses if not e.get("is_advance")})
    filter_cat = col2.selectbox(
        "📂 الفئة",
        ["الكل"] + all_cats,
    )
    date_from = col3.date_input("من", value=date.today().replace(day=1), key="rec_from")
    date_to   = col4.date_input("إلى", value=date.today(), key="rec_to")

    # Employee filter
    db = get_db()
    try:
        emps = db.query(Employee).filter(Employee.status == "Active").all()
        emp_names = ["الكل"] + [e.name for e in emps]
    finally:
        db.close()
    filter_emp = st.selectbox("👤 الموظف (للسلف)", emp_names)

    # Apply filters
    filtered = expenses[:]
    if filter_type == "سلف فقط":
        filtered = [e for e in filtered if e.get("is_advance")]
    elif filter_type == "مصروفات فقط":
        filtered = [e for e in filtered if not e.get("is_advance")]
    if filter_cat != "الكل":
        filtered = [e for e in filtered if e.get("category") == filter_cat]
    if filter_emp != "الكل":
        filtered = [e for e in filtered if e.get("emp_name") == filter_emp or e.get("desc", "").find(filter_emp) >= 0]
    filtered = [
        e for e in filtered
        if date_from <= (e["date"] if isinstance(e["date"], date) else datetime.strptime(str(e["date"]), "%Y-%m-%d").date()) <= date_to
    ]

    if not filtered:
        empty_state("لا توجد نتائج للفلاتر المحددة", "🔍")
        return

    # Summary pills
    total = sum(e["amount"] for e in filtered)
    c1, c2 = st.columns(2)
    kpi_card(c1, "📋", "عدد العمليات", str(len(filtered)), COLORS["primary"])
    kpi_card(c2, "💰", "الإجمالي",     f"{total:,.0f} ج.م", COLORS["danger"])

    st.markdown("<div style='height:12px;'></div>", unsafe_allow_html=True)

    rows = []
    for e in filtered:
        rows.append({
            "التاريخ":  str(e["date"]),
            "النوع":    "💰 سلفة" if e.get("is_advance") else "🧾 مصروف",
            "البيان":   e.get("desc", e.get("emp_name", "—")),
            "الفئة":    e.get("category", "—"),
            "الموظف":   e.get("emp_name", "—") if e.get("is_advance") else "—",
            "المبلغ":   f"{e['amount']:,.0f}",
            "ملاحظات":  e.get("notes", ""),
        })
    df = pd.DataFrame(rows)
    st.dataframe(df, use_container_width=True, hide_index=True)

    # Edit category inline (advances don't have a category, so only
    # real expense rows are eligible here)
    with st.expander("✏️ تعديل فئة مصروف", expanded=False):
        editable = [e for e in filtered if not e.get("is_advance")]
        if not editable:
            st.caption("لا توجد مصروفات (غير السلف) ضمن النتائج الحالية لتعديل فئتها.")
        else:
            idx = st.number_input("رقم المصروف (1-based, غير شامل السلف)", min_value=1,
                                  max_value=len(editable), step=1, value=1) - 1
            new_cat = st.selectbox("الفئة الجديدة", CATEGORIES, key="edit_cat_sel")
            if st.button("💾 تحديث الفئة"):
                target = editable[idx]
                db = get_db()
                try:
                    row = db.query(Expense).filter(Expense.id == target["id"]).first()
                    if row:
                        row.category = new_cat
                        db.commit()
                        success_toast("تم تحديث الفئة.")
                        st.rerun()
                    else:
                        danger_toast("لم يتم العثور على هذا المصروف، قد يكون تم حذفه.")
                finally:
                    db.close()

    if st.button("🗑️ حذف آخر مصروف", key="del_last_exp"):
        db = get_db()
        try:
            last = db.query(Expense).order_by(Expense.id.desc()).first()
            if last:
                db.delete(last)
                db.commit()
                st.rerun()
            else:
                warning_toast("لا توجد مصروفات (غير السلف) لحذفها. لحذف سلفة، استخدم صفحة «السلف».")
        finally:
            db.close()


# ─────────────────────────────────────────────────────────────────
#  TAB 5 — REPORTS & EXPORT
# ─────────────────────────────────────────────────────────────────

def _tab_reports():
    section_card("التقارير والتصدير", "📊", COLORS["info"])

    expenses = _load_ledger()
    received = _load_received_funds()

    report_type = st.selectbox("📋 نوع التقرير", [
        "ملخص يومي",
        "ملخص شهري",
        "سلف الموظفين",
        "تفصيل فئات المصروفات",
        "تقرير الرصيد",
    ])

    st.markdown("<div style='height:10px;'></div>", unsafe_allow_html=True)

    if report_type == "ملخص يومي":
        _report_daily(expenses, received)
    elif report_type == "ملخص شهري":
        _report_monthly(expenses, received)
    elif report_type == "سلف الموظفين":
        _report_advances(expenses)
    elif report_type == "تفصيل فئات المصروفات":
        _report_categories(expenses)
    else:
        _report_balance(expenses, received)

    # Export
    st.markdown("<div style='height:20px;'></div>", unsafe_allow_html=True)
    section_card("تصدير Excel", "📥", COLORS["success"])

    if st.button("📥 تصدير إلى Excel", type="primary", use_container_width=True):
        _export_excel(expenses, received)


def _report_daily(expenses, received):
    if not expenses and not received:
        empty_state("لا توجد بيانات", "📊"); return

    all_dates = sorted({str(e["date"]) for e in expenses} | {str(r["date"]) for r in received})
    rows = []
    for d in all_dates:
        day_exp  = [e for e in expenses  if str(e["date"]) == d]
        day_recv = [r for r in received  if str(r["date"]) == d]
        total_recv    = sum(r["amount"] for r in day_recv)
        total_adv     = sum(e["amount"] for e in day_exp if e.get("is_advance"))
        total_other   = sum(e["amount"] for e in day_exp if not e.get("is_advance"))
        rows.append({
            "التاريخ":       d,
            "الوارد":        f"{total_recv:,.0f}",
            "السلف":         f"{total_adv:,.0f}",
            "مصروفات أخرى": f"{total_other:,.0f}",
            "الرصيد اليومي": f"{total_recv - total_adv - total_other:,.0f}",
        })
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)


def _report_monthly(expenses, received):
    if not expenses and not received:
        empty_state("لا توجد بيانات", "📊"); return

    def month_key(d):
        dd = d if isinstance(d, date) else datetime.strptime(str(d), "%Y-%m-%d").date()
        return f"{dd.year}-{dd.month:02d}"

    all_months = sorted({month_key(e["date"]) for e in expenses} |
                        {month_key(r["date"]) for r in received})
    rows = []
    for m in all_months:
        me = [e for e in expenses if month_key(e["date"]) == m]
        mr = [r for r in received if month_key(r["date"]) == m]
        total_recv = sum(r["amount"] for r in mr)
        total_tax  = sum(r["tax"]    for r in mr)
        total_adv  = sum(e["amount"] for e in me if e.get("is_advance"))
        total_oth  = sum(e["amount"] for e in me if not e.get("is_advance"))
        rows.append({
            "الشهر":         m,
            "إجمالي الوارد": f"{total_recv:,.0f}",
            "رسوم التحويل":  f"{total_tax:,.2f}",
            "سلف الموظفين":  f"{total_adv:,.0f}",
            "مصروفات أخرى":  f"{total_oth:,.0f}",
            "الرصيد":        f"{total_recv - total_tax - total_adv - total_oth:,.0f}",
        })
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)


def _report_advances(expenses):
    advs = [e for e in expenses if e.get("is_advance")]
    if not advs:
        empty_state("لا توجد سلف مسجلة", "💰"); return
    rows = []
    by_emp = {}
    for a in advs:
        n = a.get("emp_name", "—")
        by_emp.setdefault(n, []).append(a)
    for emp, items in sorted(by_emp.items()):
        total = sum(i["amount"] for i in items)
        rows.append({
            "الموظف":    emp,
            "عدد السلف": len(items),
            "الإجمالي":  f"{total:,.0f} ج.م",
            "آخر سلفة":  str(max(i["date"] for i in items)),
        })
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)


def _report_categories(expenses):
    non_adv = [e for e in expenses if not e.get("is_advance")]
    if not non_adv:
        empty_state("لا توجد مصروفات", "🧾"); return
    by_cat = {}
    for e in non_adv:
        c = e.get("category", "أخرى")
        by_cat.setdefault(c, []).append(e["amount"])
    total_all = sum(e["amount"] for e in non_adv)
    rows = []
    for cat, amts in sorted(by_cat.items(), key=lambda x: -sum(x[1])):
        s = sum(amts)
        rows.append({
            "الفئة":       cat,
            "عدد البنود":  len(amts),
            "الإجمالي":    f"{s:,.0f} ج.م",
            "النسبة":      f"{s/total_all*100:.1f}%" if total_all else "0%",
        })
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)


def _report_balance(expenses, received):
    total_recv = sum(r["amount"] for r in received)
    total_tax  = sum(r["tax"]    for r in received)
    net_recv   = total_recv - total_tax
    total_adv  = sum(e["amount"] for e in expenses if e.get("is_advance"))
    total_exp  = sum(e["amount"] for e in expenses if not e.get("is_advance"))
    balance    = net_recv - total_adv - total_exp

    bal_color  = COLORS["success"] if balance >= 0 else COLORS["danger"]

    st.markdown(f"""
<div style="background:{COLORS['surface']};border:1px solid {COLORS['border']};
            border-radius:14px;padding:22px 28px;margin:10px 0;">
  <table style="width:100%;border-collapse:collapse;font-size:14px;">
    <tr style="border-bottom:1px solid {COLORS['border']};"><td style="padding:10px 0;color:{COLORS['text_muted']};">إجمالي الوارد</td><td style="text-align:left;font-weight:700;color:{COLORS['success']};">+ {total_recv:,.0f} ج.م</td></tr>
    <tr style="border-bottom:1px solid {COLORS['border']};"><td style="padding:10px 0;color:{COLORS['text_muted']};">رسوم التحويل</td><td style="text-align:left;font-weight:700;color:{COLORS['warning']};"> − {total_tax:,.2f} ج.م</td></tr>
    <tr style="border-bottom:2px solid {COLORS['primary']};"><td style="padding:10px 0;color:{COLORS['primary']};font-weight:700;">صافي الوارد</td><td style="text-align:left;font-weight:800;color:{COLORS['primary']};">{net_recv:,.2f} ج.م</td></tr>
    <tr style="border-bottom:1px solid {COLORS['border']};"><td style="padding:10px 0;color:{COLORS['text_muted']};">إجمالي السلف</td><td style="text-align:left;font-weight:700;color:{COLORS['danger']};"> − {total_adv:,.0f} ج.م</td></tr>
    <tr style="border-bottom:1px solid {COLORS['border']};"><td style="padding:10px 0;color:{COLORS['text_muted']};">مصروفات أخرى</td><td style="text-align:left;font-weight:700;color:{COLORS['danger']};"> − {total_exp:,.0f} ج.م</td></tr>
    <tr><td style="padding:14px 0;font-size:18px;font-weight:800;color:{bal_color};">الرصيد المتبقي</td><td style="text-align:left;font-size:18px;font-weight:900;color:{bal_color};">{balance:,.0f} ج.م</td></tr>
  </table>
</div>
""", unsafe_allow_html=True)


def _export_excel(expenses, received):
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        st.error("مكتبة openpyxl غير مثبتة. شغّل: pip install openpyxl")
        return

    wb = openpyxl.Workbook()

    # Sheet 1: Received Funds
    ws1 = wb.active
    ws1.title = "الواردات"
    hdr1 = ["التاريخ", "المبلغ المستلم", "رسوم التحويل",
             "الصافي", "المصدر", "ملاحظات"]
    ws1.append(hdr1)
    for r in received:
        ws1.append([str(r["date"]), r["amount"], r["tax"], r["net"],
                    r["source"], r["notes"]])

    # Sheet 2: Expenses
    ws2 = wb.create_sheet("المصروفات")
    ws2.append(["التاريخ", "البيان", "الفئة", "النوع", "الموظف", "المبلغ", "ملاحظات"])
    for e in expenses:
        ws2.append([
            str(e["date"]),
            e.get("desc", e.get("emp_name", "")),
            e.get("category", "—"),
            "سلفة" if e.get("is_advance") else "مصروف",
            e.get("emp_name", "—") if e.get("is_advance") else "—",
            e["amount"],
            e.get("notes", ""),
        ])

    # Sheet 3: Advances
    ws3 = wb.create_sheet("سلف الموظفين")
    ws3.append(["التاريخ", "الموظف", "المبلغ", "ملاحظات"])
    for e in [x for x in expenses if x.get("is_advance")]:
        ws3.append([str(e["date"]), e.get("emp_name", "—"),
                    e["amount"], e.get("notes", "")])

    # Sheet 4: Summary
    ws4 = wb.create_sheet("الملخص")
    total_recv = sum(r["amount"] for r in received)
    total_tax  = sum(r["tax"]    for r in received)
    net_recv   = total_recv - total_tax
    total_adv  = sum(e["amount"] for e in expenses if e.get("is_advance"))
    total_exp  = sum(e["amount"] for e in expenses if not e.get("is_advance"))
    balance    = net_recv - total_adv - total_exp
    ws4.append(["البند", "المبلغ"])
    for row in [
        ("إجمالي الوارد", total_recv),
        ("رسوم التحويل",  total_tax),
        ("صافي الوارد",   net_recv),
        ("إجمالي السلف",  total_adv),
        ("مصروفات أخرى",  total_exp),
        ("الرصيد المتبقي", balance),
    ]:
        ws4.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    st.download_button(
        label="⬇️ تحميل ملف Excel",
        data=buf.getvalue(),
        file_name=f"expenses_petty_cash_{date.today()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    success_toast("تم إنشاء ملف Excel بنجاح!")
